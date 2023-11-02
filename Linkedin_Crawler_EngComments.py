from selenium import webdriver
import time
import math
import openpyxl

# Step 1: Initializing the WebDriver
# Creating the 'browser' variable, the 'webdriver.Chrome()' command opens a Chrome browser window.
browser = webdriver.Chrome()

# Access the LinkedIn website using the '.get("url")' command.
browser.get("https://www.linkedin.com")
# Using 'time.sleep' to give time for the web page to load.
time.sleep(2)
# Logging into the LinkedIn website.
browser.find_element_by_name("session_key").send_keys("<e-mail>") # Your LinkedIn email
time.sleep(2)
browser.find_element_by_name("session_password").send_keys("********") # Your LinkedIn password
time.sleep(2)
browser.find_element_by_id("login-submit").click() # Click to finish the login

# Step 1: Initializing the WebDriver

# Step 2: Searching for Companies According to Description/Sector
while True:
    try:
        search = input("Enter the industry to search on LinkedIn: ")
        # Access the page with search results.
        browser.get('https://www.linkedin.com/search/results/companies/?keywords=' + search + '&origin=SWITCH_SEARCH_VERTICAL')
        time.sleep(2)

        # Capturing the number of results (pages). This command stores the element on the page that contains the search result numbers in the 'result_elem' variable.
        result_elem = browser.find_element_by_xpath("//h3[@class='search-results__total t-14 t-black--light t-normal pl5 pt4 clear-both']")

    except:
        print("No companies found, please try again.")
        continue

    else:
        break

# Extracting the result text and converting it to a number.
result = result_elem.get_attribute("innerText")
result = result.replace(',', '')
companies = [int(s) for s in result.split() if s.isdigit]
pages = int(math.ceil(companies[0] / 10))

# Step 2: Searching for Companies According to Description/Sector

# Step 3: Loop Through the Search Results, Page by Page
p = 1

while p <= pages:

    print("Page", p, "of", pages) # Page counter

    # Scroll the page to capture elements properly.
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/8)")
    time.sleep(1)
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/4)")
    time.sleep(1)
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/1)")
    elem_url = browser.find_elements_by_xpath("//a[@class='search-result__result-link ember-view']")

    N_url = len(elem_url)
    url = [""] * N_url

    # Loop through the companies on the page.
    for k in range(0, N_url):
        url[k] = elem_url[k].get_attribute('href')

    # Visit each company's page.
    for i in range(0, N_url, 2):
        browser.get(url[i])
        time.sleep(1)
        
        # Extract various information about the company.
        try:
            elem_nome = browser.find_elements_by_xpath("//h1[@class='org-top-card-module__name t-24 t-black t-light']")
            nome = elem_nome[0].get_attribute('innerText')
        except:
            nome = ""

        try:
            elem_size = browser.find_elements_by_xpath("//p[@class='org-about-company-module__company-staff-count-range t-14 t-black--light t-normal mb3']")
            size = elem_size[0].get_attribute('innerText')
        except:
            size = ""

        try:
            elem_site = browser.find_elements_by_xpath("//a[@class='org-about-us-company-module__website mb3 link-without-visited-state ember-view']")
            site = elem_site[0].get_attribute('innerText')
        except:
            site = ""

        try:
            elem_special = browser.find_elements_by_xpath("//p[@class='org-about-company-module__specialities mb5 t-14 t-black--light t-normal mb3']")
            special = elem_special[0].get_attribute('innerText')
        except:
            special = ""

        try:
            elem_head = browser.find_elements_by_xpath("//p[@class='org-about-company-module__headquarters t-14 t-black--light t-normal mb3']")
            head = elem_head[0].get_attribute('innerText')
        except:
            head = ""

        book = openpyxl.load_workbook('Linkedin_Crawler.xlsx')
        sheet = book.active

        if p == 1:
            sheet.cell(row=i/2+3, column=1).value = nome
            sheet.cell(row=i/2+3, column=2).value = url[i]
            sheet.cell(row=i/2+3, column=3).value = site
            sheet.cell(row=i/2+3, column=4).value = special
            sheet.cell(row=i/2+3, column=5).value = size
            sheet.cell(row=i/2+3, column=6).value = head
        else:
            sheet.cell(row=i/2+3+10*(p-1), column=1).value = nome
            sheet.cell(row=i/2+3+10*(p-1), column=2).value = url[i]
            sheet.cell(row=i/2+3+10*(p-1), column=3).value = site
            sheet.cell(row=i/2+3+10*(p-1), column=4).value = special
            sheet.cell(row=i/2+3+10*(p-1), column=5).value = size
            sheet.cell(row=i/2+3+10*(p-1), column=6).value = head

        try:
            # Attempt to fetch information about the company's CEOs, Directors, or Directoras.
            # Unfortunately, LinkedIn has recently restricted the use of additional parameters for search.
            page_elem = browser.find_elements_by_xpath("//a[@class='org-company-employees-snackbar__details-highlight snackbar-description-see-all-link link-without-visited-state ember-view']")
            page = page_elem[0].get_attribute("href")

            url_emp = page + '&keywords=(CEO)%20OR%20(diretor)%20OR%20(diretora)&origin=GLOBAL_SEARCH_HEADER'
            browser.get(url_emp)
            time.sleep(1)
            
            # Copy employee names.
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/8)")
            time.sleep(1)
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/4)")
            time.sleep(1)
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/1)")

            employees = browser.find_elements_by_xpath("//span[@class='name-and-icon']")
            N_emp = len(employees)
            link_emp_elm = browser.find_elements_by_xpath("//a[@data-control-name='search_srp_result']")
            cargo_emp_elm = browser.find_elements_by_xpath("//p[@class='subline-level-1 t-14 t-black t-normal search-result__truncate']")

            for j in range(0, N_emp):
                nome_emp = employees[j].get_attribute('innerText')
                nome_emp = nome_emp.split("  ")[0]
                cargo_emp = cargo_emp_elm[j].get_attribute('innerText')
                link_emp = link_emp_elm[2*j].get_attribute('href')

                if p == 1:
                    sheet.cell(row=i/2+3, column=3*j+7).value = nome_emp
                    sheet.cell(row=i/2+3, column=3*j+8).value = cargo_emp
                    sheet.cell(row=i/2+3, column=3*j+9).value = link_emp
                else:
                    sheet.cell(row=i/2+3+10*(p-1), column=3*j+7).value = nome_emp
                    sheet.cell(row=i/2+3+10*(p-1), column=3*j+8).value = cargo_emp
                    sheet.cell(row=i/2+3+10*(p-1), column=3*j+9).value = link_emp

        except:
            pass

        book.save('Linkedin_Crawler.xlsx')

    p += 1
    browser.get('https://www.linkedin.com/search/results/companies/?keywords='+search+'&origin=SWITCH_SEARCH_VERTICAL&page='+ str(p))
