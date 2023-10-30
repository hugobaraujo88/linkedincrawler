from selenium import webdriver
import time
import math
import openpyxl


####### 1) Início: Inicializando o webdriver

#Criando a variável browser, comando webdriver.Chrome() abre uma janela do chrome.
browser = webdriver.Chrome()

#Acessa o site do linkedin, comando .get("url").
browser.get("https://www.linkedin.com")
#Toda vez que aparece time.sleep, significa que o compilador vai dar um tempo antes de executar o próximo comando
#É importante porque as vezes a internet demora para carregar
time.sleep(2)
#Fazendo o login no site do linkedin.
browser.find_element_by_name("session_key").send_keys("hugoaraujob@gmail.com") #Email no linkedin
time.sleep(2)
browser.find_element_by_name("session_password").send_keys("********") #Senha no linkedin
time.sleep(2)
browser.find_element_by_id("login-submit").click() #clica para finalizar o login

####### 1) Fim: Inicializando o webdriver


####### 2) Início: Busca de empresas de acordo com a descrição/setor

while True:
    try:
        busca = input("Digite o segemento para busca no linkedin: ")
        #Acessa a página com resultados da busca
        browser.get('https://www.linkedin.com/search/results/companies/?keywords='+busca+'&origin=SWITCH_SEARCH_VERTICAL')
        time.sleep(2)

        #Capturando o número de resultados (paginas). Esse comando armazena o elemento da página que contem os resultados da busca na
        #variável "result_elem". Por exemplo, quando você digita "software gestão" vai aparecer "Showing 5,727 results".
        #o "result_elem" vai armazenar o elemento de "html" que contem o texto "Showing 5,727 results". O elemento em html será nesse caso:
        #"<h3 class="search-results__total t-14 t-black--light t-normal pl5 pt4 clear-both"> Showing 5,727 results </h3>"
        result_elem = browser.find_element_by_xpath("//h3[@class='search-results__total t-14 t-black--light t-normal pl5 pt4 clear-both']")


    except:
        print("Nenhuma empresa encontrada, tente novamente")
        continue

    else:
        break

#Esse comando extrai um atributo do elemento armazenado na varável "result_elem" e armazena na variável "result". No caso é um texto por isso o "innerText".
#No exemplo acima seria o texto "Showing 5,727 results" que está dentro do elemento
result = result_elem.get_attribute("innerText")

#Série de comandos para transformar o texto extraído em número

#O texto está na forma "Showing 5,727 results", o que o comando faz é retirar a vírgula do "5,727". A variável se torna "result = Showing 5727 results".
result = result.replace(',','')
#Retira os caracteres númericos da string "result = Showing 5727 results" e armazena na variável "empresas", i.e. empresas = [5727] ainda é um string.
empresas = [int(s) for s in result.split() if s.isdigit()]
#Calcula o número de páginas: Converte a string "empresas" em inteiro pela função int(). Divide por 10 e arredonda pra cima (math.ceil) pois há no máximo 10 empresas por página. Nesse caso "paginas=573"
paginas = int(math.ceil(empresas[0]/10))

####### 2) Fim: Busca de empresas de acordo com a descrição/setor


####### 3) Início do loop nos resultados da busca, página por página

#Agora o algoritmo vai buscar as empresas página por página começando da página 1. Nesse caso, vai ser um loop de p=1 até p=páginas (no exemplo p=1 até p=573, página 1 até a 573).
p = 1

while p <= paginas:

    #Contador de quantas páginas já foram exploradas
    print("página",p,"de",paginas)

    #comando para rolar a página do início até o fim, isso se mostrou necessário para os comandos que capturam elementos das páginas.
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/8)")
    time.sleep(1)
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/4)")
    time.sleep(1)
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight/1)")

    #Armazena na variável "elem_url" os elementos que contém as links das empresas no linkedin
    elem_url = browser.find_elements_by_xpath("//a[@class='search-result__result-link ember-view']")

    #A variável "elem_url" é um vetor. Esse comando armazena o tamanho do vetor na variável "N_url" para fazer o loop das empresas na página.
    N_url = len(elem_url)
    url = [""]*N_url

    ####### 3.1) Início do loop na página p, empresa por empresa

    for k in range(0 , N_url):

        #Esse comando extrai um atributo do elemento armazenado na varável "elem_url" e armazena na variável "url". No caso é um link por isso o "href". O link da empresa
        url[k] = elem_url[k].get_attribute('href')

    #Visitando empresa por empresa. Está indo de 2 em 2 porque o elemento //a[@class='search-result__result-link ember-view'] por algum motivo está duplicado.
    #Por exemplo a Empresa A aparece 2x consecutivamente na posição url[0} e url[1], então o correto é ir url[0], url[2], url[4]...
    for i in range(0, N_url, 2):

        browser.get(url[i])
        time.sleep(1)

        #Pegando todos as características disponíveis da empresa nome, faixa de funcionários, site da empresa, descrição das especialidades, cidade sede

        #Pegar nome da empresa
        try:
            elem_nome = browser.find_elements_by_xpath("//h1[@class='org-top-card-module__name t-24 t-black t-light']")
            nome = elem_nome[0].get_attribute('innerText')
        except:
            nome = ""

        #Pegar faixa de funcionários
        try:
            elem_size = browser.find_elements_by_xpath("//p[@class='org-about-company-module__company-staff-count-range t-14 t-black--light t-normal mb3']")
            size = elem_size[0].get_attribute('innerText')
        except:
            size = ""

        #Pegar site se houver
        try:
            elem_site = browser.find_elements_by_xpath("//a[@class='org-about-us-company-module__website mb3 link-without-visited-state ember-view']")
            site = elem_site[0].get_attribute('innerText')
        except:
            size = ""

        #Pegar descrição das especialidades se houver
        try:
            elem_special=browser.find_elements_by_xpath("//p[@class='org-about-company-module__specialities mb5 t-14 t-black--light t-normal mb3']")
            special=elem_special[0].get_attribute('innerText')
        except:
            special = ""

        #Pegar Headquarters
        try:
            elem_head=browser.find_elements_by_xpath("//p[@class='org-about-company-module__headquarters t-14 t-black--light t-normal mb3']")
            head=elem_head[0].get_attribute('innerText')
        except:
            head = ""

        #Procedimento para escrever os dados na planilha de excel
        book = openpyxl.load_workbook('Linkedin_Crawler.xlsx')
        sheet = book.active
        if p==1:
            sheet.cell(row=i/2+3, column=1).value = nome
            sheet.cell(row=i/2+3, column=2).value = url[i]
            sheet.cell(row=i/2+3, column=3).value = site
            sheet.cell(row=i/2+3, column=4).value = special
            sheet.cell(row=i/2+3, column=5).value = size
            sheet.cell(row=i/2+3, column=6).value = head
        else:
            sheet.cell(row=i/2+3+10*(p-1), column=1).value = nome
            sheet.cell(row=i/2+3+10*(p-1), column=2).value = url[i]
            sheet.cell(row=i/2+3+10*(p-1), column=3).value = site
            sheet.cell(row=i/2+3+10*(p-1), column=4).value = special
            sheet.cell(row=i/2+3+10*(p-1), column=5).value = size
            sheet.cell(row=i/2+3+10*(p-1), column=6).value = head

        ####### 3.1.2) Início da tentativa de pegar informações sobre os sócios/diretores


        try:

            #Comando que armazena o elemento que contém o link da página de funcionários da empresa na variável "pagina_socios_elem". É o link "See all XX employees on Linkedin"
            pagina_socios_elm = browser.find_elements_by_xpath("//a[@class='org-company-employees-snackbar__details-highlight snackbar-description-see-all-link link-without-visited-state ember-view']")

            #Esse comando extrai um atributo do elemento armazenado na varável "pagina_socios_elm" e armazena na variável "pagina_socios". No caso é um link por isso o "href".
            pagina_socios = pagina_socios_elm[0].get_attribute("href")

            #A partir da página de funcionários da empresa, tenta buscar CEO ou Diretor ou Diretora. Linkedin deixava usar mais parâmetros para busca, infelizmente eles bloquearam recentemente essa possibilidade.
            url_emp = pagina_socios + '&keywords=(CEO)%20OR%20(diretor)%20OR%20(diretora)&origin=GLOBAL_SEARCH_HEADER'
            browser.get(url_emp)
            time.sleep(1)

            #Copiar o nome das pessoas

            #Comando para rolar a página do início até o fim, isso se mostrou necessário para os comandos que capturam elementos das páginas.
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/8)")
            time.sleep(1)
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/4)")
            time.sleep(1)
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight/1)")

            employees = browser.find_elements_by_xpath("//span[@class='name-and-icon']") #Armazena o elemento que tem o nome dos funcionários no vetor 'employees' / nao duplicado
            N_emp = len(employees) #Coloca na varável N_emp o tamanho do vetor 'employees'
            link_emp_elm = browser.find_elements_by_xpath("//a[@data-control-name='search_srp_result']") #Armazena o o elemento que tem os links na variável link_emp_elm/ duplicado
            cargo_emp_elm = browser.find_elements_by_xpath("//p[@class='subline-level-1 t-14 t-black t-normal search-result__truncate']") #Armazena o elemento que tem o cargo

            #Loop para armazenar todos os proprietários/diretores na planilha de excel Nome, cargo e linkedin da pessoa.
            for j in range (0,N_emp):
                nome_emp = employees[j].get_attribute('innerText')
                nome_emp = nome_emp.split("  ")[0]
                cargo_emp= cargo_emp_elm[j].get_attribute('innerText')
                link_emp = link_emp_elm[2*j].get_attribute('href')

                if p==1:
                    sheet.cell(row=i/2+3,column=3*j+7).value = nome_emp
                    sheet.cell(row=i/2+3,column=3*j+8).value = cargo_emp
                    sheet.cell(row=i/2+3,column=3*j+9).value = link_emp
                else:
                    sheet.cell(row=i/2+3+10*(p-1),column=3*j+7).value = nome_emp
                    sheet.cell(row=i/2+3+10*(p-1),column=3*j+8).value = cargo_emp
                    sheet.cell(row=i/2+3+10*(p-1),column=3*j+9).value = link_emp

        except:
            pass

        ####### 3.1.2) Fim da tentativa de pegar informações sobre os sócios/diretores

        book.save('Linkedin_Crawler.xlsx')

    ####### 3.1) Fim do loop na página p, empresa por empresa

####### 3) Fim do loop nos resultados da busca, página por página

    p += 1
    browser.get('https://www.linkedin.com/search/results/companies/?keywords='+busca+'&origin=SWITCH_SEARCH_VERTICAL&page='+ str(p))


